# ========= Copyright 2023-2024 @ CAMEL-AI.org. All Rights Reserved. =========
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ========= Copyright 2023-2024 @ CAMEL-AI.org. All Rights Reserved. =========

# ruff: noqa

from collections.abc import Callable

import pandas as pd

from utu.tools import AsyncBaseToolkit
from utu.utils import get_logger

logger = get_logger(__name__)


class ExcelToolkit(AsyncBaseToolkit):
    def _convert_to_markdown(self, df: pd.DataFrame) -> str:
        r"""Convert DataFrame to Markdown format table.

        Args:
            df (pd.DataFrame): DataFrame containing the Excel data.

        Returns:
            str: Markdown formatted table.
        """
        from tabulate import tabulate

        md_table = tabulate(df, headers="keys", tablefmt="pipe")
        return str(md_table)

    async def extract_excel_content(self, document_path: str, max_char_length=5000, return_cell_info=False) -> str:
        r"""Extract detailed cell information from an Excel file, including
        multiple sheets.

        Args:
            document_path (str): The path of the Excel file.
            max_char_length (int): The maximum character length for each sheet's content.
            return_cell_info (bool): Whether to return detailed cell information. Defaults to False. Unless the font color or fill color info is needed, it is recommended to set it to False to reduce the length of the returned content.

        Returns:
            str: Extracted excel information, including details of each sheet.
        """
        from openpyxl import load_workbook
        from xls2xlsx import XLS2XLSX

        logger.debug(f"Calling extract_excel_content with document_path: {document_path}")

        if not (document_path.endswith("xls") or document_path.endswith("xlsx") or document_path.endswith("csv")):
            logger.error("Only xls, xlsx, csv files are supported.")
            return f"Failed to process file {document_path}: It is not excel format. Please try other ways."

        if document_path.endswith("csv"):
            try:
                df = pd.read_csv(document_path)
                md_table = self._convert_to_markdown(df)
                return f"CSV File Processed:\n{md_table}"
            except Exception as e:
                logger.error(f"Failed to process file {document_path}: {e}")
                return f"Failed to process file {document_path}: {e}"

        if document_path.endswith("xls"):
            output_path = document_path.replace(".xls", ".xlsx")
            x2x = XLS2XLSX(document_path)
            x2x.to_xlsx(output_path)
            document_path = output_path

        # Load the Excel workbook
        wb = load_workbook(document_path, data_only=True)
        sheet_info_list = []

        # Iterate through all sheets
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            cell_info_list = []

            for row in ws.iter_rows():
                for cell in row:
                    row_num = cell.row
                    col_letter = cell.column_letter

                    cell_value = cell.value

                    font_color = None
                    if cell.font and cell.font.color and "rgb=None" not in str(cell.font.color):  # Handle font color
                        font_color = cell.font.color.rgb

                    fill_color = None
                    if (
                        cell.fill and cell.fill.fgColor and "rgb=None" not in str(cell.fill.fgColor)
                    ):  # Handle fill color
                        fill_color = cell.fill.fgColor.rgb

                    cell_info_list.append(
                        {
                            "index": f"{row_num}{col_letter}",
                            "value": cell_value,
                            "font_color": font_color,
                            "fill_color": fill_color,
                        }
                    )

            # Convert the sheet to a DataFrame and then to markdown
            sheet_df = pd.read_excel(document_path, sheet_name=sheet, engine="openpyxl")
            markdown_content = self._convert_to_markdown(sheet_df)

            # Collect all information for the sheet
            sheet_info = {
                "sheet_name": sheet,
                "cell_info_list": cell_info_list,
                "markdown_content": markdown_content,
            }
            sheet_info_list.append(sheet_info)

        # if sheet_info is too long, only return the first n characters
        MAX_CHAR_LENGTH = int(max_char_length)
        result_str = ""
        for sheet_info in sheet_info_list:
            cell_info = str(sheet_info["cell_info_list"])
            markdown_content = str(sheet_info["markdown_content"])

            if len(cell_info) > MAX_CHAR_LENGTH:
                cell_info = cell_info[:MAX_CHAR_LENGTH]
                cell_info = cell_info + f"... (Truncated, total length is {len(cell_info)})"
            if len(markdown_content) > MAX_CHAR_LENGTH:
                markdown_content = markdown_content[:MAX_CHAR_LENGTH]
                markdown_content = (
                    markdown_content
                    + f"... (Truncated, total length is {len(markdown_content)}, please write python code to get the full content)"
                )

            if return_cell_info:
                result_str += f"""
                Sheet Name: {sheet_info["sheet_name"]}
                Cell information list:
                {cell_info}
                
                Markdown View of the content:
                {markdown_content}
                
                {"-" * 40}
                """
            else:
                result_str += f"""
                Sheet Name: {sheet_info["sheet_name"]}
                
                Markdown View of the content:
                {markdown_content}
                
                {"-" * 40}
                """

        return result_str

    async def get_tools_map(self) -> dict[str, Callable]:
        return {
            "extract_excel_content": self.extract_excel_content,
        }
